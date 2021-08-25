from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import NoSuchElementException
import time
import pandas as pd
from openpyxl import load_workbook
import json


class DuoObj:
    def __init__(self, text, index):
        self.text = text
        self.index = index

def performance(fn):
    def wrapper(*args, **kwargs):
        t1 = time.time()
        result = fn(*args, **kwargs)
        t2 = time.time()
        print('\n')
        print(f'function took {t2-t1} s')
        print('\n')
        return result
    return wrapper


def click_continue(browser_drv, delay_time):
    try:
        myElem = WebDriverWait(browser_drv, delay_time).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, '.Ejd2j')))
        browser_drv.find_element_by_css_selector(".Ejd2j").click()
    except NoSuchElementException:
        pass
    except TimeoutException:
        pass
    except StaleElementReferenceException:
        pass


def load_story(browser_drv, delay_time, story_link, story_name):
    browser_drv.get(story_link)
    try:
        myElem = WebDriverWait(browser_drv, delay_time).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, '.QmLbM')))
        print('\n')
        print("Story " + story_name + " has been loaded")
        driver.find_element_by_css_selector('.QmLbM').click()
    except TimeoutException:
        print("Loading took too much time!")


def two_options_click( btn_xpath, btn_text, text_loc):
    global driver
    global delay
    str(btn_xpath)
    str(btn_text)
    str(text_loc)
    try:
        xpath_one = btn_xpath + '1]/button'
        xpath_two = btn_xpath + '2]/button'
        text_loc_one = text_loc + '1]/div/div/span[1]'
        text_loc_two = text_loc + '2]/div/div/span[1]'

        First_option = driver.find_element_by_xpath(text_loc_one)
        Second_option = driver.find_element_by_xpath(text_loc_two)

        if First_option.text == btn_text:
            option_one = driver.find_elements_by_xpath(xpath_one)[0]
            option_one.click()
        elif Second_option.text == btn_text:
            option_two = driver.find_elements_by_xpath(xpath_two)[0]
            option_two.click()
    except (NoSuchElementException, StaleElementReferenceException, TimeoutException):
        pass

def three_options_click(btn_text, location_xpath, relative_xpath):
    global driver
    global delay
    str(location_xpath)
    str(relative_xpath)
    str(btn_text)
    try:
        myElem = WebDriverWait(driver, delay).until(EC.element_to_be_clickable(
            (By.XPATH, location_xpath + '3]/button')))

        First_option = driver.find_element_by_xpath(location_xpath + "1" + relative_xpath)
        Second_option = driver.find_element_by_xpath(location_xpath + "2" + relative_xpath)
        Third_option = driver.find_element_by_xpath(location_xpath + "3" + relative_xpath)

        if First_option.text == btn_text:
            driver.find_elements_by_xpath(location_xpath + "1]/button")[0].click()
        elif Second_option.text == btn_text:
            driver.find_elements_by_xpath(location_xpath + "2]/button")[0].click()
        elif Third_option.text == btn_text:
            driver.find_elements_by_xpath(location_xpath + "3]/button")[0].click()
    except (NoSuchElementException, StaleElementReferenceException, TimeoutException):
        pass

def clicktext(buttonloc, *args, **kwargs):
    myElem = WebDriverWait(driver, delay).until(EC.element_to_be_clickable(
            (By.XPATH, str(buttonloc) + "span[1]")))
    for arg in args:
        driver.find_elements_by_xpath(str(buttonloc) + "*[@class='_37HGt'][text()='" + arg + "']")[0].click()

def matchpairs(jsonpath, xpathtobutton, xlsxpath):
    global driver
    str(xpathtobutton)
    str(jsonpath)
    words = {}

    FirstSlot = driver.find_element_by_xpath(xpathtobutton + '1]/button')
    words[FirstSlot.text] = "1"

    SecondSlot = driver.find_element_by_xpath(xpathtobutton + '2]/button')
    words[SecondSlot.text] = "2"

    ThirdSlot = driver.find_element_by_xpath(xpathtobutton + '3]/button')
    words[ThirdSlot.text] = "3"

    ForthSlot = driver.find_element_by_xpath(xpathtobutton + '4]/button')
    words[ForthSlot.text] = "4"

    FirthSlot = driver.find_element_by_xpath(xpathtobutton + '5]/button')
    words[FirthSlot.text] = "5"

    SixthSlot = driver.find_element_by_xpath(xpathtobutton + '6]/button')
    words[SixthSlot.text] = "6"

    SeventhSlot = driver.find_element_by_xpath(xpathtobutton + '7]/button')
    words[SeventhSlot.text] = "7"

    EighthSlot = driver.find_element_by_xpath(xpathtobutton + '8]/button')
    words[EighthSlot.text] = "8"

    NinethSlot = driver.find_element_by_xpath(xpathtobutton + '9]/button')
    words[NinethSlot.text] = "9"

    TenthSlot = driver.find_element_by_xpath(xpathtobutton + '10]/button')
    words[TenthSlot.text] = "10"

    print(words)

    df = pd.DataFrame(data=words, index=[0])

    df = (df.T)
    print(df)

    filename = str(xlsxpath)
    book = load_workbook(filename)
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}

    for sheetname in writer.sheets:
        df.to_excel(writer, sheet_name=sheetname,
                    startrow=writer.sheets[sheetname].max_row, index=True, header=False)

    writer.save()

    with open(jsonpath, "r", encoding="utf8") as json_file:
        filename1 = json.load(json_file)

    dict2 = filename1 

    list1 = []
    list2 = []
    list3 = []
    list4 = []
    list5 = []
    list6 = []
    list7 = []
    list8 = []
    list9 = []

    for item in dict2:
        if dict2[item] == FirstSlot.text:
            list1.append(item)

    len(list1)
    if (len(list1) > 1):
        try:
            FirstPairMatch = words.get(list1[0])
            driver.find_element_by_xpath(
                xpathtobutton + str(FirstPairMatch) + ']/button').click()
            FirstSlot.click()
        except (TypeError, NoSuchElementException):
            FirstPairMatch = words.get(list1[1])
            driver.find_element_by_xpath(
                xpathtobutton + str(FirstPairMatch) + ']/button').click()
            FirstSlot.click()
        except (TypeError, NoSuchElementException):
            FirstPairMatch = words.get(list1[2])
            driver.find_element_by_xpath(
                xpathtobutton + str(FirstPairMatch) + ']/button').click()
            FirstSlot.click()
    else:
        FirstPair = dict2.get(FirstSlot.text)
        print(FirstPair)
        FirstPairMatch = words.get(FirstPair)
        driver.find_element_by_xpath(
            xpathtobutton + str(FirstPairMatch) + ']/button').click()
        FirstSlot.click()

    if (SecondSlot.is_displayed() and SecondSlot.is_enabled()):
        for item in dict2:
            if dict2[item] == SecondSlot.text:
                list2.append(item)
        len(list2)
        if (len(list2) > 1):
            try:
                SecondPairMatch = words.get(list2[0])
                driver.find_element_by_xpath(
                    xpathtobutton + str(SecondPairMatch) + ']/button').click()
                SecondSlot.click()
            except (TypeError, NoSuchElementException):
                SecondPairMatch = words.get(list2[1])
                driver.find_element_by_xpath(
                    xpathtobutton + str(SecondPairMatch) + ']/button').click()
                SecondSlot.click()
            except (TypeError, NoSuchElementException):
                SecondPairMatch = words.get(list2[2])
                driver.find_element_by_xpath(
                    xpathtobutton + str(SecondPairMatch) + ']/button').click()
                SecondSlot.click()
        else:
            SecondPair = dict2.get(SecondSlot.text)
            print(SecondPair)
            SecondPairMatch = words.get(SecondPair)
            driver.find_element_by_xpath(
                xpathtobutton + str(SecondPairMatch) + ']/button').click()
            SecondSlot.click()

    if (ThirdSlot.is_displayed() and ThirdSlot.is_enabled()):
        for item in dict2:
            if dict2[item] == ThirdSlot.text:
                list3.append(item)
        len(list3)
        if (len(list3) > 1):
            try:
                ThirdPairMatch = words.get(list3[0])
                driver.find_element_by_xpath(
                    xpathtobutton + str(ThirdPairMatch) + ']/button').click()
                ThirdSlot.click()
            except (TypeError, NoSuchElementException):
                ThirdPairMatch = words.get(list3[1])
                driver.find_element_by_xpath(
                    xpathtobutton + str(ThirdPairMatch) + ']/button').click()
                ThirdSlot.click()
            except (TypeError, NoSuchElementException):
                ThirdPairMatch = words.get(list3[2])
                driver.find_element_by_xpath(
                    xpathtobutton + str(ThirdPairMatch) + ']/button').click()
                ThirdSlot.click()
        else:
            ThirdPair = dict2.get(ThirdSlot.text)
            print(ThirdPair)
            ThirdPairMatch = words.get(ThirdPair)
            driver.find_element_by_xpath(
                xpathtobutton + str(ThirdPairMatch) + ']/button').click()
            ThirdSlot.click()


    if (ForthSlot.is_displayed() and ForthSlot.is_enabled()):
        for item in dict2:
            if dict2[item] == ForthSlot.text:
                list4.append(item)
        len(list4)
        if (len(list4) > 1):
            try:
                ForthPairMatch = words.get(list4[0])
                driver.find_element_by_xpath(
                    xpathtobutton + str(ForthPairMatch) + ']/button').click()
                ForthSlot.click()
            except (TypeError, NoSuchElementException):
                ForthPairMatch = words.get(list4[1])
                driver.find_element_by_xpath(
                    xpathtobutton + str(ForthPairMatch) + ']/button').click()
                ForthSlot.click()
            except (TypeError, NoSuchElementException):
                ForthPairMatch = words.get(list4[2])
                driver.find_element_by_xpath(
                    xpathtobutton + str(ForthPairMatch) + ']/button').click()
                ForthSlot.click()
        else:
            ForthPair = dict2.get(ForthSlot.text)
            print(ForthPair)
            ForthPairMatch = words.get(ForthPair)
            driver.find_element_by_xpath(
                xpathtobutton + str(ForthPairMatch) + ']/button').click()
            ForthSlot.click()

    if (FirthSlot.is_displayed() and FirthSlot.is_enabled()):
        for item in dict2:
            if dict2[item] == FirthSlot.text:
                list5.append(item)
        len(list5)
        if (len(list5) > 1):
            try:
                FirthPairMatch = words.get(list5[0])
                driver.find_element_by_xpath(
                    xpathtobutton + str(FirthPairMatch) + ']/button').click()
                FirthSlot.click()
            except (TypeError, NoSuchElementException):
                FirthPairMatch = words.get(list5[1])
                driver.find_element_by_xpath(
                    xpathtobutton + str(FirthPairMatch) + ']/button').click()
                FirthSlot.click()
            except (TypeError, NoSuchElementException):
                FirthPairMatch = words.get(list5[2])
                driver.find_element_by_xpath(
                    xpathtobutton + str(FirthPairMatch) + ']/button').click()
                FirthSlot.click()
        else:
            FirthPair = dict2.get(FirthSlot.text)
            print(FirthPair)
            FirthPairMatch = words.get(FirthPair)
            driver.find_element_by_xpath(
                xpathtobutton + str(FirthPairMatch) + ']/button').click()
            FirthSlot.click()

    if (SixthSlot.is_displayed() and SixthSlot.is_enabled()):
        for item in dict2:
            if dict2[item] == SixthSlot.text:
                list6.append(item)
        len(list6)
        if (len(list6) > 1):
            try:
                SixthPairMatch = words.get(list6[0])
                driver.find_element_by_xpath(
                    xpathtobutton + str(SixthPairMatch) + ']/button').click()
                SixthSlot.click()
            except (TypeError, NoSuchElementException):
                SixthPairMatch = words.get(list6[1])
                driver.find_element_by_xpath(
                    xpathtobutton + str(SixthPairMatch) + ']/button').click()
                SixthSlot.click()
            except (TypeError, NoSuchElementException):
                SixthPairMatch = words.get(list6[2])
                driver.find_element_by_xpath(
                    xpathtobutton + str(SixthPairMatch) + ']/button').click()
                SixthSlot.click()
        else:
            SixthPair = dict2.get(SixthSlot.text)
            print(SixthPair)
            SixthPairMatch = words.get(SixthPair)
            driver.find_element_by_xpath(
                xpathtobutton + str(SixthPairMatch) + ']/button').click()
            SixthSlot.click()

    if (SeventhSlot.is_displayed() and SeventhSlot.is_enabled()):
        for item in dict2:
            if dict2[item] == SeventhSlot.text:
                list7.append(item)
        len(list7)
        if (len(list7) > 1):
            try:
                SeventhPairMatch = words.get(list7[0])
                driver.find_element_by_xpath(
                    xpathtobutton + str(SeventhPairMatch) + ']/button').click()
                SeventhSlot.click()
            except (TypeError, NoSuchElementException):
                SeventhPairMatch = words.get(list7[1])
                driver.find_element_by_xpath(
                    xpathtobutton + str(SeventhPairMatch) + ']/button').click()
                SeventhSlot.click()
            except (TypeError, NoSuchElementException):
                SeventhPairMatch = words.get(list7[2])
                driver.find_element_by_xpath(
                    xpathtobutton + str(SeventhPairMatch) + ']/button').click()
                SeventhSlot.click()
        else:
            SeventhPair = dict2.get(SeventhSlot.text)
            print(SeventhPair)
            SeventhPairMatch = words.get(SeventhPair)
            driver.find_element_by_xpath(
                xpathtobutton + str(SeventhPairMatch) + ']/button').click()
            SeventhSlot.click()

    if (EighthSlot.is_displayed() and EighthSlot.is_enabled()):
        for item in dict2:
            if dict2[item] == EighthSlot.text:
                list8.append(item)
        len(list8)
        if (len(list8) > 1):
            try:
                EighthPairMatch = words.get(list8[0])
                driver.find_element_by_xpath(
                    xpathtobutton + str(EighthPairMatch) + ']/button').click()
                EighthSlot.click()
            except (TypeError, NoSuchElementException):
                EighthPairMatch = words.get(list8[1])
                driver.find_element_by_xpath(
                    xpathtobutton + str(EighthPairMatch) + ']/button').click()
                EighthSlot.click()
            except (TypeError, NoSuchElementException):
                EighthPairMatch = words.get(list8[2])
                driver.find_element_by_xpath(
                    xpathtobutton + str(EighthPairMatch) + ']/button').click()
                EighthSlot.click()
        else:
            EighthPair = dict2.get(EighthSlot.text)
            print(EighthPair)
            EighthPairMatch = words.get(EighthPair)
            driver.find_element_by_xpath(
                xpathtobutton + str(EighthPairMatch) + ']/button').click()
            EighthSlot.click()

    if (NinethSlot.is_displayed() and NinethSlot.is_enabled()):
        for item in dict2:
            if dict2[item] == NinethSlot.text:
                list9.append(item)
        len(list9)
        if (len(list9) > 1):
            try:
                NinethPairMatch = words.get(list9[0])
                driver.find_element_by_xpath(
                    xpathtobutton + str(NinethPairMatch) + ']/button').click()
                NinethSlot.click()
            except (TypeError, NoSuchElementException):
                NinethPairMatch = words.get(list9[1])
                driver.find_element_by_xpath(
                    xpathtobutton + str(NinethPairMatch) + ']/button').click()
                NinethSlot.click()
            except (TypeError, NoSuchElementException):
                NinethPairMatch = words.get(list9[2])
                driver.find_element_by_xpath(
                    xpathtobutton + str(NinethPairMatch) + ']/button').click()
                NinethSlot.click()
        else:
            NinethPair = dict2.get(NinethSlot.text)
            print(NinethPair)
            NinethPairMatch = words.get(NinethPair)
            driver.find_element_by_xpath(
                xpathtobutton + str(NinethPairMatch) + ']/button').click()
            NinethSlot.click()
    # Continue
    click_continue(driver, delay)

    myElem = WebDriverWait(driver, delay).until(EC.element_to_be_clickable(
        (By.XPATH, '/html/body/div[1]/div/div/div/div/div[2]/div/div/div/button')))
    driver.find_element_by_xpath(
        "/html/body/div[1]/div/div/div/div/div[2]/div/div/div/button").click()
    try:
        myElem = WebDriverWait(driver, delay).until(EC.element_to_be_clickable(
            (By.XPATH, '/html/body/div[1]/div/div/div/div/div[2]/div/div/div/button')))
        driver.find_element_by_xpath(
            "html/body/div[1]/div/div/div/div/div[3]/div/div[2]/div/button").click()            
    except (NoSuchElementException, StaleElementReferenceException, TimeoutException):
        pass

driver = webdriver.Edge(executable_path= >>>>Path to edge chrominium selenium driver<<<<)

driver.set_page_load_timeout(10)
driver.get("https://www.duolingo.com/")
delay = 6  # seconds
try:
    myElem = WebDriverWait(driver, delay).until(EC.element_to_be_clickable(
        (By.XPATH, '//*[@id="root"]/div/div/span[1]/div/div[1]/div[2]/div[2]/a')))
    driver.find_element_by_xpath(
        '/html/body/div[1]/div/div/span[1]/div/div[1]/div[2]/div[2]/a').click()
    # .presence_of_element_located
except TimeoutException:
    print("Loading took too much time!")

driver.find_element_by_css_selector(
    "div._2a3s4:nth-child(1) > label:nth-child(1) > div:nth-child(1) > input:nth-child(1)").send_keys(>>>your email<<<)
driver.find_element_by_css_selector(
    "div._2a3s4:nth-child(2) > label:nth-child(1) > div:nth-child(1) > input:nth-child(1)").send_keys(>>>>your password<<<<)

try:
    myElem = WebDriverWait(driver, delay).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, '._2oW4v')))
    driver.find_element_by_css_selector('._2oW4v').click()
except TimeoutException:
    print("Loading took too much time!")

time.sleep(2)

@performance
def story_good_morning(mode='listen'):
    global driver
    global delay
    # Load and start Story
    # Load Good Morning! story
    if mode == 'listen':
        load_story(driver, delay, "https://www.duolingo.com/stories/de-guten-morgen?mode=listen", "Guten Morgen listen mode")
    elif mode == 'read':
        load_story(driver, delay, "https://www.duolingo.com/stories/de-guten-morgen?mode=read", "Guten Morgen read mode")

    # Continue
    click_continue(driver, delay)

    # Rainer und Anna sind zu Hause
    if mode == 'listen':
        clicktext('/html/body/div[1]/div/div/div/div/div[1]/div[1]/div[4]/div/div/', 'Frau', 'Anna', 'sind', 'zu', 'Hause')

    click_continue(driver, delay)

    # Rainer is at home with his wife
    if mode == 'read': 
        try:
            First_option = driver.find_elements_by_xpath(
                "/html/body/div[1]/div[1]/div/div/div/div[1]/div[1]/div[3]/div/ul/li[1]/div/div/span[1][@class='_2igzU _3LUrt _2P5W7'][contains(text(),'Yes')]")
            Second_option = driver.find_elements_by_xpath(
                "/html/body/div[1]/div[1]/div/div/div/div[1]/div[1]/div[3]/div/ul/li[2]/div/div/span[1][@class='_2igzU _3LUrt _2P5W7'][contains(text(),'Yes')]")

            if First_option:
                driver.find_elements_by_xpath(
                    "//*[@id='root']/div[1]/div/div/div/div[1]/div[1]/div[3]/div/ul/li[1]/button")[0].click()
            elif Second_option:
                driver.find_elements_by_xpath(
                    "//*[@id='root']/div[1]/div/div/div/div[1]/div[1]/div[3]/div/ul/li[2]/button")[0].click()
        except (NoSuchElementException, StaleElementReferenceException, TimeoutException):
            pass

    # Continue
    click_continue(driver, delay)

    # Guten Morgen Anna
    click_continue(driver, delay)

    # Guten Morgen
    click_continue(driver, delay)

    # Wo ist mein Englischbuch?
    if mode == 'listen':
        clicktext('/html/body/div[1]/div/div/div/div/div[1]/div[1]/div[9]/div/div/', 'Wo', 'ist', 'mein', 'Englischbuch')

    click_continue(driver, delay)

    # Anna wants to know
    if mode == 'read':
        three_options_click('where', "/html/body/div[1]/div/div/div/div/div[1]/div[1]/div[7]/div/ul/li[", ']/div/div/span[2]')
    
    # Continue
    click_continue(driver, delay)

    # Dein Englischbuch
    click_continue(driver, delay)

    # Ja mein Englischbuch fur die Universitat
    click_continue(driver, delay)

    # Ich brauche mein Buch
    click_continue(driver, delay)

    # Anna Dein buch ist hier
    click_continue(driver, delay)

    #Danke, Rainer
    click_continue(driver, delay)

    # Tut mir leid ich bin mude
    click_continue(driver, delay)

    # Mochtest du eine Kaffee
    if mode == 'listen':
        clicktext('/html/body/div[1]/div/div/div/div/div[1]/div[1]/div[18]/div/div/', 'Möchtest', 'du', 'einen', 'Kaffee')

    click_continue(driver, delay)

    if mode == 'read':
        three_options_click('coffee', "//*[@id='root']/div/div/div/div/div[1]/div[1]/div[15]/div/ul/li[", ']/div/div/span[6]')

    # Continue
    click_continue(driver, delay)

    # Ja bitte
    click_continue(driver, delay)

    # Milch
    click_continue(driver, delay)

    # Nein danke
    click_continue(driver, delay)

    # Here ist dein Kaffe
    click_continue(driver, delay)

    # Anna gibt Zucker in den Kaffee
    if mode == 'listen':
        clicktext('/html/body/div[1]/div/div/div/div/div[1]/div[1]/div[25]/div/div/', 'in', 'den', 'Kaffee')

    click_continue(driver, delay)

    # What is Anna doing?
    if mode == 'read':
        try:
            myElem = WebDriverWait(driver, delay).until(EC.element_to_be_clickable(
                (By.XPATH, '/html/body/div[1]/div/div/div/div/div[1]/div[1]/div[21]/div/ul/li[3]/button')))
        except (NoSuchElementException, StaleElementReferenceException, TimeoutException):
            pass

        try:
            First_option = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div/div/div/div[1]/div[1]/div[21]/div/ul/li[1]/div/div/span[13]")
        except NoSuchElementException:
            First_option = DuoObj(False, 1)
            First_option.text = 'False'

        try:
            Second_option = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div/div/div/div[1]/div[1]/div[21]/div/ul/li[2]/div/div/span[13]")
        except NoSuchElementException:
            Second_option = DuoObj(False, 2)
            Second_option.text = 'False'

        try:
            Third_option = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div/div/div/div[1]/div[1]/div[21]/div/ul/li[3]/div/div/span[13]")
        except NoSuchElementException:
            Third_option = DuoObj(False, 3)
            Third_option.text = 'False'

        try:
            First_option2 = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div/div/div/div[1]/div[1]/div[21]/div/ul/li[1]/div/div/span[5]")
        except NoSuchElementException:
            First_option2 = DuoObj(False, 1)
            First_option2.text = 'False'

        try:
            Second_option2 = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div/div/div/div[1]/div[1]/div[21]/div/ul/li[2]/div/div/span[5]")
        except NoSuchElementException:
            Second_option2 = DuoObj(False, 2)
            Second_option2.text = 'False'

        try:
            Third_option2 = driver.find_element_by_xpath(
                "/html/body/div[1]/div/div/div/div/div[1]/div[1]/div[21]/div/ul/li[3]/div/div/span[5]")
        except NoSuchElementException:
            Third_option2 = DuoObj(False, 3)
            Third_option2.text = 'False'

        if (First_option.text == "coffee" and First_option2.text == "putting"):
            driver.find_elements_by_xpath(
                "/html/body/div[1]/div/div/div/div/div[1]/div[1]/div[21]/div/ul/li[1]/button")[0].click()
        elif (Second_option.text == "coffee" and Second_option2.text == "putting"):
            driver.find_elements_by_xpath(
                "/html/body/div[1]/div/div/div/div/div[1]/div[1]/div[21]/div/ul/li[2]/button")[0].click()
        elif (Third_option.text == "coffee" and Third_option2.text == "putting"):
            driver.find_elements_by_xpath(
                "/html/body/div[1]/div/div/div/div/div[1]/div[1]/div[21]/div/ul/li[3]/button")[0].click()

    # Continue
    click_continue(driver, delay)

    # Sie trinkt den Kaffee
    # Continue
    click_continue(driver, delay)

    # Igitt!
    # Continue
    click_continue(driver, delay)

    # was?
    # Continue
    click_continue(driver, delay)

    # What does 'was' mean?
    if mode == 'read':
        three_options_click("What", '/html/body/div[1]/div/div/div/div/div[1]/div[1]/div[25]/div/ul/li[', ']/div/div/span[1]')


    # Continue
    click_continue(driver, delay)


    #das ist Saltz
    click_continue(driver, delay)

    # Anna du bist nicht mude du bist total mude
    click_continue(driver, delay)


    # Du brauchst viel Kaffee
    if mode == 'listen':
        clicktext('/html/body/div[1]/div/div/div/div/div[1]/div[1]/div[33]/div/div/', 'Du', 'brauchst', 'viel', 'Kaffee')

    click_continue(driver, delay)

    if mode == 'read':
        three_options_click("salt", '/html/body/div[1]/div/div/div/div/div[1]/div[1]/div[29]/div/ul/li[', ']/div/div/span[6]')


    # Continue
    click_continue(driver, delay)
    time.sleep(1)
    if mode == 'listen':
        matchpairs(>>>>Path to json dictionary with correct pairs <<<<, '/html/body/div[1]/div/div/div/div/div[1]/div[1]/div[34]/div/ul/li[', >>>>Path to xlsx file collector for all occuring words to manually pair them correctly for dictionary<<<<)
    elif mode == 'read':
        matchpairs(>>>>Path to json dictionary with correct pairs <<<<, '/html/body/div[1]/div/div/div/div/div[1]/div[1]/div[30]/div/ul/li[', >>>>Path to xlsx file collector for all occuring words to manually pair them correctly for dictionary<<<<)


    matchpairs(>>>>Path to json dictionary with correct pairs <<<<, '/html/body/div[1]/div/div/div/div/div[1]/div[1]/div[34]/div/ul/li[', >>>>Path to xlsx file collector for all occuring words to manually pair them correctly for dictionary<<<<)

#call story function in listen mode
story_good_morning('listen')
#call story function in read mode
story_good_morning('read')
